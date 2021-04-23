import requests, openpyxl, time, os
from bs4 import BeautifulSoup
from pathlib import Path
from openpyxl import load_workbook
from datetime import datetime
from openpyxl.styles import Font, Fill
from openpyxl.utils import get_column_letter

#Modificar para tomar otro jugador -- Here you set the player account
usuario = 'Prometeo'
nroBattleTag = '11597'

#función borrar pantalla -- Cleans Screen
borrarPantalla = lambda: os.system ("cls")

#Control de archivo Excel -- Excel File control. If not exists, creates it
archivo = 'Rangos '+ usuario + '.xlsx'
fileObj = Path(archivo)
if fileObj.is_file() == False:
    wb = openpyxl.Workbook()
    wb.save(archivo)
    ws = wb.active
    ws.title = "TANQUE"
    ws.append(('Fecha', 'Hora', 'Valor'))
    ws['A1'].font = Font(bold=True)
    ws['B1'].font = Font(bold=True)    
    ws['C1'].font = Font(bold=True)
    ws['G1'].font = Font(bold=True)
    ws['J1'].font = Font(bold=True)
    ws["G1"] = "Últ. Valor: "
    ws["J1"] = "Últ. Control: "
    for i in range(1, ws.max_column+1):
        ws.column_dimensions[get_column_letter(i)].bestFit = True
        ws.column_dimensions[get_column_letter(i)].auto_size = True
    ws = wb.create_sheet('DAÑO')
    ws.append(('Fecha', 'Hora', 'Valor'))
    ws['A1'].font = Font(bold=True)
    ws['B1'].font = Font(bold=True)
    ws['C1'].font = Font(bold=True)
    ws['G1'].font = Font(bold=True)
    ws['J1'].font = Font(bold=True)
    ws["G1"] = "Últ. Valor: "
    ws["J1"] = "Últ. Control: "
    for i in range(1, ws.max_column+1):
        ws.column_dimensions[get_column_letter(i)].bestFit = True
        ws.column_dimensions[get_column_letter(i)].auto_size = True
    ws = wb.create_sheet('SOPORTE')
    ws.append(('Fecha', 'Hora', 'Valor'))
    ws['A1'].font = Font(bold=True)
    ws['B1'].font = Font(bold=True)
    ws['C1'].font = Font(bold=True)
    ws['G1'].font = Font(bold=True)
    ws['J1'].font = Font(bold=True)
    ws["G1"] = "Últ. Valor: "
    ws["J1"] = "Últ. Control: "
    for i in range(1, ws.max_column+1):
        ws.column_dimensions[get_column_letter(i)].bestFit = True
        ws.column_dimensions[get_column_letter(i)].auto_size = True
    wb.save(archivo)

wb = load_workbook(filename = archivo)

#Para que se ejecute constantemente -- So it loops forever
while 1 == 1:

    #Uso de requests 
    URL = 'https://playoverwatch.com/es-es/career/pc/'+usuario+'-'+nroBattleTag+'/'
    page = requests.get(URL)

    #Uso de BeautifulSoup
    soup = BeautifulSoup(page.content, 'html.parser')
    results = soup.find(id='overview-section')

    #Filtrado de datos - Data Filtering
    desc_sr = results.find_all("div", {"class": "competitive-rank-tier"})
    sr = results.find_all("div", {"class": "competitive-rank-level"})
    longitudCadena = (len(sr)/2)
    sr_clean = sr[:int(longitudCadena)]


    #Presentación de datos -- Data Presentation
    for s in sr_clean:
        if desc_sr[sr_clean.index(s)]['data-ow-tooltip-text'].upper().replace("ÍNDICE DE HABILIDAD: ", "") == "TANQUE":
            wb.active = 0
            ws = wb.active
            ultimoValor = ws["H1"].value
        elif desc_sr[sr_clean.index(s)]['data-ow-tooltip-text'].upper().replace("ÍNDICE DE HABILIDAD: ", "") == "DAÑO":
            wb.active = 1
            ws = wb.active
            ultimoValor = ws["H1"].value
        else:
            wb.active = 2
            ws = wb.active
            ultimoValor = ws["H1"].value
                
        ws["K1"].value = datetime.now()

        if str(ultimoValor) != s.text.strip():
            hoja = wb.active
            ws["H1"].value = int(s.text.strip())
            
            datos = [str(datetime.now().strftime("%d/%m/%y")),str(datetime.now().strftime('%H:%M:%S')),int(s.text.strip())]

            hoja.append(datos)    
    
    wb.save(archivo)
    
    #Información para el usuario -- State of process info
    borrarPantalla() 
    print("Se puede cerrar el programa..")
    print("Último Control: " + datetime.now().strftime("%d/%m/%y %H:%M:%S"))
    time.sleep(120)
    borrarPantalla() 
    print("Procesando... no cerrar")
    